import os
import glob
import json
import sys
from datetime import datetime, timezone, timedelta

from scraper import fetch_news, fetch_company_news
from analyzer import analyze_news
from formatter import format_email
from notifier import send_email

IST = timezone(timedelta(hours=5, minutes=30))


def _find_xlsx() -> str | None:
    matches = glob.glob("*.xlsx") + glob.glob("*.XLSX")
    return matches[0] if matches else None


def _load_from_excel(path: str) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    stocks = []
    for row in rows[1:]:
        if not any(row):
            continue
        d = dict(zip(headers, row))
        symbol = str(d.get("symbol", "") or "").strip().upper()
        sector = str(d.get("sector", "") or "").strip()
        if symbol:
            stocks.append({"symbol": symbol, "sector": sector})
    return stocks


def load_watchlist() -> list:
    xlsx = _find_xlsx()
    if xlsx:
        stocks = _load_from_excel(xlsx)
        print(f"  Loaded {len(stocks)} stocks from {xlsx}")
        return stocks
    with open("watchlist.json") as f:
        stocks = json.load(f)["stocks"]
    print(f"  Loaded {len(stocks)} stocks from watchlist.json")
    return stocks


def _fill_watchlist_gaps(analysis: dict, watchlist: list) -> dict:
    """Add NOT IMPACTED for any watchlist stocks Claude omitted."""
    summary = analysis.setdefault("watchlist_summary", {})
    for stock in watchlist:
        if stock["symbol"] not in summary:
            summary[stock["symbol"]] = {
                "sentiment": "NOT IMPACTED",
                "trigger": "No relevant news today",
            }
    return analysis


def main():
    now = datetime.now(tz=IST)
    print(f"[{now.strftime('%Y-%m-%d %H:%M IST')}] Morning news synthesis starting...")

    watchlist = load_watchlist()
    if not watchlist:
        print("ERROR: Watchlist is empty. Exiting.")
        sys.exit(1)

    print("Fetching macro news from Moneycontrol and Economic Times...")
    macro_articles = fetch_news(hours_back=24)
    if not macro_articles:
        print("ERROR: No macro articles fetched. Exiting.")
        sys.exit(1)

    print("Fetching company-specific news for watchlist stocks...")
    company_articles = fetch_company_news(watchlist, hours_back=24)

    articles = macro_articles + company_articles
    print(f"Total: {len(articles)} articles ({len(macro_articles)} macro + {len(company_articles)} company-specific)")
    print(f"Analyzing {len(articles)} articles with Claude...")
    analysis = analyze_news(articles, watchlist)
    analysis = _fill_watchlist_gaps(analysis, watchlist)

    date_str = now.strftime("%A, %d %B %Y")
    subject = f"Morning Market Brief — {date_str}"
    html_body = format_email(analysis, date_str)

    recipient = os.environ["RECIPIENT_EMAIL"]
    send_email(subject, html_body, recipient)
    print(f"Done. Email sent to {recipient}")


if __name__ == "__main__":
    main()
